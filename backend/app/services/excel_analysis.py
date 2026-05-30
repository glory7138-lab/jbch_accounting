from __future__ import annotations

import re
from collections import Counter
from datetime import date as date_type, datetime as dt_type
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session
import uuid

from app.models import Account, Fund, ImportBatch, Member, Voucher

KEYWORDS = {
    "date": ["날짜", "일자", "거래일", "년월일", "입금일자", "date"],
    "description": ["적요", "내역", "내용", "품목", "description", "memo"],
    "income": ["수입", "입금", "입", "차변", "income", "debit", "헌금"],
    "expense": ["지출", "출금", "출", "대변", "expense", "credit"],
    "amount": ["금액", "합계", "잔액", "총액", "amount", "balance"],
    "account": ["계정", "계정과목", "과목", "코드", "account", "회계코드"],
    "category": ["구분", "분류", "관리항목", "category", "type"],
    "note": ["비고", "참고", "remark", "note"],
    "name": ["이름", "성명", "name"],
    "member_no": ["번호", "no", "id"],
}

ACCOUNT_CODE_CANDIDATES = ["회계코드", "계정코드"]
MEMBER_SHEET_CANDIDATES = ["헌금봉투번호", "헌금 봉투 번호", "성도현황번호", "성도 현황 번호", "번호"]
FUND_EXCLUDE_NAMES = {"sheet3", "계정코드", "작성상의 주의사항", "회계결산 표지", "sheet1"}


def normalize(value: Any) -> str:
    if value is None or pd.isna(value):
        return ""
    text = str(value).strip()
    text = re.sub(r"\s+", " ", text)
    return "" if text.lower() == "nan" else text


def classify_header(header: str) -> list[str]:
    lowered = normalize(header).lower()
    hits: list[str] = []
    for label, keywords in KEYWORDS.items():
        if any(keyword.lower() in lowered for keyword in keywords):
            hits.append(label)
    return hits


def find_header_row(df: pd.DataFrame) -> int:
    best_idx = 0
    best_score = -1.0
    for idx in range(min(len(df), 12)):
        row = [normalize(v) for v in df.iloc[idx].tolist()]
        non_empty = [v for v in row if v and v.lower() != "nan"]
        if not non_empty:
            continue
        score = 0.0
        for cell in non_empty:
            score += 3 if classify_header(cell) else 0
            score += 1 if len(cell) <= 20 else 0
        score += min(len(non_empty), 12)
        if score > best_score:
            best_idx = idx
            best_score = score
    return best_idx


def load_sheet_frame(path: Path, sheet_name: str) -> tuple[pd.DataFrame, int]:
    raw = pd.read_excel(path, sheet_name=sheet_name, header=None)
    if raw.empty:
        return raw, 0
    header_idx = find_header_row(raw)
    header_row = [normalize(v) for v in raw.iloc[header_idx].tolist()]

    headers: list[str] = []
    seen: Counter[str] = Counter()
    for i, header in enumerate(header_row):
        value = header if header and header.lower() != "nan" else f"unnamed_{i + 1}"
        seen[value] += 1
        if seen[value] > 1:
            value = f"{value}_{seen[value]}"
        headers.append(value)

    data = raw.iloc[header_idx + 1 :].copy()
    data.columns = headers
    data = data.dropna(how="all")
    keep_columns = []
    for col in data.columns:
        if not str(col).startswith("unnamed_") or data[col].notna().any():
            keep_columns.append(col)
    data = data.loc[:, keep_columns]
    return data, header_idx + 1


def analyze_workbook(path: Path) -> dict[str, Any]:
    # openpyxl only supports .xlsx; for .xls we skip merged cell info
    merged_info: dict[str, list[str]] = {}
    if path.suffix.lower() == ".xlsx":
        try:
            workbook = load_workbook(path, data_only=True)
            for sn in workbook.sheetnames:
                merged_info[sn] = [str(rng) for rng in workbook[sn].merged_cells.ranges][:20]
            workbook.close()
        except Exception:
            pass

    excel = pd.ExcelFile(path)
    result: dict[str, Any] = {"file_name": path.name, "sheet_names": excel.sheet_names, "sheets": []}
    for sheet_name in excel.sheet_names:
        data, header_row_index = load_sheet_frame(path, sheet_name)
        headers = [str(col) for col in data.columns.tolist()]
        sample_rows = []
        for _, row in data.head(5).iterrows():
            sample = {str(k): normalize(v) for k, v in row.items() if normalize(v)}
            if sample:
                sample_rows.append(sample)

        normalized_headers = [{"header": h, "tags": classify_header(h)} for h in headers]
        role_tags = Counter(tag for item in normalized_headers for tag in item["tags"])
        if role_tags["income"] and role_tags["expense"]:
            role_guess = "journal_or_summary"
        elif role_tags["amount"] and role_tags["date"]:
            role_guess = "cash_or_balance"
        elif role_tags["category"] and role_tags["amount"]:
            role_guess = "summary"
        else:
            role_guess = "ledger_like"

        result["sheets"].append(
            {
                "sheet_name": sheet_name,
                "header_row_index": header_row_index,
                "headers": headers,
                "normalized_headers": normalized_headers,
                "sample_rows": sample_rows,
                "non_empty_rows": int(len(data)),
                "merged_ranges": merged_info.get(sheet_name, []),
                "role_guess": role_guess,
            }
        )
    return result


def build_schema_brief(analysis: list[dict[str, Any]]) -> dict[str, Any]:
    files_summary = []
    header_bank: list[str] = []
    for workbook in analysis:
        files_summary.append(
            {
                "file_name": workbook["file_name"],
                "sheet_count": len(workbook["sheet_names"]),
                "sheet_names": workbook["sheet_names"],
            }
        )
        for sheet in workbook["sheets"]:
            header_bank.extend(sheet["headers"])

    inferred_domains = [
        {"table": "funds", "reason": "회계장부 파일에 통합계정, 일반계정, 부서/헌금별 시트가 반복됨"},
        {"table": "accounts", "reason": "계정코드 시트에 회계코드와 관리항목 계층이 존재함"},
        {"table": "members", "reason": "헌금현황 파일에 헌금 봉투 번호 시트가 있어 봉투번호가 사실상 식별키 역할을 함"},
        {"table": "vouchers", "reason": "회계장부 시트들이 날짜, 적요, 입금/지출, 잔액 중심의 거래 원장을 가짐"},
        {"table": "voucher_lines", "reason": "수입/지출과 계정코드가 함께 존재해 다중 분개 확장이 필요함"},
        {"table": "accounting_periods", "reason": "월말결산 양식과 주간보고 자료가 월 단위 마감을 전제로 함"},
        {"table": "import_batches", "reason": "엑셀 파싱 결과를 추적하고 재가져오기 이력을 관리해야 함"},
        {"table": "ai_suggestion_logs", "reason": "적요 기반 계정추천 기록과 개선 피드백이 필요함"},
    ]

    return {
        "files": files_summary,
        "top_headers": sorted(set([header for header in header_bank if header and not header.startswith("unnamed")]))[:80],
        "inferred_domains": inferred_domains,
    }


def _list_workbooks(base_dir: str | Path) -> list[Path]:
    base_path = Path(base_dir)
    results: list[Path] = []
    # Search root and found/ subdirectory for both .xlsx and .xls
    search_dirs = [base_path]
    found_dir = base_path / "found"
    if found_dir.is_dir():
        search_dirs.append(found_dir)
    for search_dir in search_dirs:
        for ext in ("*.xlsx", "*.xls"):
            for path in search_dir.glob(ext):
                if not path.name.startswith("~$"):
                    results.append(path)
    # Deduplicate by filename (prefer found/ over root)
    seen_names: set[str] = set()
    unique: list[Path] = []
    for path in sorted(results, key=lambda p: (0 if "found" in str(p) else 1, p.name)):
        if path.name not in seen_names:
            seen_names.add(path.name)
            unique.append(path)
    return sorted(unique)


def analyze_sample_directory(base_dir: str | Path) -> dict[str, Any]:
    workbooks = _list_workbooks(base_dir)
    analysis = [analyze_workbook(path) for path in workbooks]
    return {"files": analysis, "schema_brief": build_schema_brief(analysis)}


def _find_sheet_name(sheet_names: list[str], candidates: list[str]) -> str | None:
    for name in sheet_names:
        lowered = name.replace(" ", "")
        if any(candidate.replace(" ", "") in lowered for candidate in candidates):
            return name
    return None


def _match_column(columns: list[str], keywords: list[str]) -> str | None:
    lowered_map = {col: normalize(col).replace(" ", "").lower() for col in columns}
    for keyword in keywords:
        target = keyword.replace(" ", "").lower()
        for col, lowered in lowered_map.items():
            if target in lowered:
                return col
    return None


def seed_reference_data(db: Session, base_dir: str | Path) -> dict[str, int]:
    workbooks = _list_workbooks(base_dir)
    imported = 0

    for workbook_path in workbooks:
        exists = db.scalar(select(ImportBatch).where(ImportBatch.source_name == workbook_path.name))
        if not exists:
            summary = analyze_workbook(workbook_path)
            db.add(ImportBatch(source_name=workbook_path.name, summary_json=summary))
            imported += 1

    ledger_workbook = next((path for path in workbooks if "회계장부" in path.name), None)
    cash_workbook = next((path for path in workbooks if "현금현황" in path.name or "헌금현황" in path.name), None)

    funds_created = 0
    if ledger_workbook:
        excel = pd.ExcelFile(ledger_workbook)
        for sheet_name in excel.sheet_names:
            normalized_name = sheet_name.strip().lower()
            if normalized_name in FUND_EXCLUDE_NAMES:
                continue
            existing_fund = db.scalar(select(Fund).where(Fund.name == sheet_name))
            if existing_fund:
                continue
            code = re.sub(r"[^0-9A-Za-z가-힣]+", "-", sheet_name).strip("-").lower() or f"fund-{funds_created + 1}"
            db.add(Fund(code=code, name=sheet_name))
            funds_created += 1

    accounts_created = 0
    accounts_updated = 0
    if ledger_workbook:
        excel = pd.ExcelFile(ledger_workbook)
        account_sheet = _find_sheet_name(excel.sheet_names, ACCOUNT_CODE_CANDIDATES)
        if account_sheet:
            data, _ = load_sheet_frame(ledger_workbook, account_sheet)
            columns = [str(c) for c in data.columns]
            code_col = _match_column(columns, ["회계코드", "계정코드"])
            major_col = _match_column(columns, ["대분류관리항목", "대분류"])
            middle_col = _match_column(columns, ["중분류관리항목", "중분류"])
            report_col = _match_column(columns, ["세부계정항목", "세부관리항목", "보고용", "예산안"])
            type_col = _match_column(columns, ["계정유형"])
            finance_col = _match_column(columns, ["재정구분"])
            debit_col = _match_column(columns, ["차변계정", "차변"])
            credit_col = _match_column(columns, ["대변계정", "대변"])
            name_col = middle_col or major_col or report_col
            for _, row in data.iterrows():
                code = normalize(row.get(code_col)) if code_col else ""
                if not code or code.lower() == "nan":
                    continue
                name = normalize(row.get(name_col)) if name_col else code
                account_payload = {
                    "name": name or code,
                    "major_category": normalize(row.get(major_col)) if major_col else None,
                    "middle_category": normalize(row.get(middle_col)) if middle_col else None,
                    "report_category": normalize(row.get(report_col)) if report_col else None,
                    "account_type": normalize(row.get(type_col)) if type_col else None,
                    "finance_category": normalize(row.get(finance_col)) if finance_col else None,
                    "debit_account": normalize(row.get(debit_col)) if debit_col else None,
                    "credit_account": normalize(row.get(credit_col)) if credit_col else None,
                }
                normal_side = None
                if account_payload["debit_account"] and not account_payload["credit_account"]:
                    normal_side = "debit"
                elif account_payload["credit_account"] and not account_payload["debit_account"]:
                    normal_side = "credit"
                account_payload["normal_side"] = normal_side

                existing_account = db.scalar(select(Account).where(Account.code == code))
                if existing_account:
                    changed = False
                    for field, value in account_payload.items():
                        if getattr(existing_account, field) != value:
                            setattr(existing_account, field, value)
                            changed = True
                    if changed:
                        accounts_updated += 1
                    continue

                db.add(Account(code=code, **account_payload))
                accounts_created += 1

    members_created = 0
    if cash_workbook:
        excel = pd.ExcelFile(cash_workbook)
        member_sheet = _find_sheet_name(excel.sheet_names, MEMBER_SHEET_CANDIDATES)
        if member_sheet:
            # 시트명에서 연도 숫자 파싱 (예: "헌금 봉투 번호 (2026)" -> 2026)
            year_match = re.search(r"\d{4}", member_sheet)
            sheet_year = int(year_match.group(0)) if year_match else 2026

            data, _ = load_sheet_frame(cash_workbook, member_sheet)
            columns = [str(c) for c in data.columns]
            no_col = _match_column(columns, ["번호", "no"])
            name_col = _match_column(columns, ["이름", "성명"])
            dept_col = _match_column(columns, ["회", "부서"])
            district_col = _match_column(columns, ["구역"])
            age_col = _match_column(columns, ["나이", "연령", "반"])
            gender_col = _match_column(columns, ["남", "여", "성별", "구분"])
            if name_col:
                for _, row in data.iterrows():
                    name = normalize(row.get(name_col))
                    if not name or name.lower() == "nan":
                        continue
                    member_no = normalize(row.get(no_col)) if no_col else None
                    if not member_no:
                        continue

                    # 이름이 같은 기존 성도의 person_id 조회
                    existing_person_id = db.scalar(
                        select(Member.person_id)
                        .where(Member.name == name)
                        .limit(1)
                    )
                    
                    if existing_person_id:
                        person_id = existing_person_id
                    else:
                        person_id = f"P-{uuid.uuid4().hex[:8].upper()}"

                    member_payload = {
                        "name": name,
                        "department_name": normalize(row.get(dept_col)) if dept_col else None,
                        "district_name": normalize(row.get(district_col)) if district_col else None,
                        "gender_or_section": normalize(row.get(gender_col)) if gender_col else None,
                        "age_or_class": normalize(row.get(age_col)) if age_col else None,
                        "source_sheet": member_sheet,
                    }

                    # 해당 연도에 이 봉투번호를 사용하는 멤버가 있는지 확인
                    exists = db.scalar(
                        select(Member)
                        .where(Member.year == sheet_year, Member.member_no == member_no)
                    )
                    if exists:
                        # 해당 연도 레코드 업데이트
                        exists.person_id = person_id
                        for field, value in member_payload.items():
                            if getattr(exists, field) != value:
                                setattr(exists, field, value)
                        continue

                    db.add(
                        Member(
                            person_id=person_id,
                            year=sheet_year,
                            member_no=member_no,
                            **member_payload,
                        )
                    )
                    members_created += 1

    # ---- Seed vouchers from 회계장부 ledger sheets ----
    LEDGER_SHEETS = [
        "통합계정", "일반계정", "교회학교후원회비", "사랑의헌금",
        "선교회비", "건축계정", "승강기계정", "해외후원", "국내선교",
    ]
    MONTH_MAP = {
        "1월": 1, "2월": 2, "3월": 3, "4월": 4, "5월": 5, "6월": 6,
        "7월": 7, "8월": 8, "9월": 9, "10월": 10, "11월": 11, "12월": 12,
    }
    WEEK_MAP = {
        "1주차": 1, "2주차": 2, "3주차": 3, "4주차": 4, "5주차": 5,
    }

    vouchers_created = 0
    if ledger_workbook:
        # Check if vouchers already exist for this workbook
        existing_voucher = db.scalar(
            select(Voucher).where(Voucher.source_workbook == ledger_workbook.name).limit(1)
        )
        if not existing_voucher:
            excel = pd.ExcelFile(ledger_workbook)
            # Determine year from file name (e.g. "2025년_회계장부_0108.xls")
            year_match = re.search(r"(\d{4})년", ledger_workbook.name)
            ledger_year = int(year_match.group(1)) if year_match else 2025

            for sheet_name in excel.sheet_names:
                if sheet_name not in LEDGER_SHEETS:
                    continue
                try:
                    data, _ = load_sheet_frame(ledger_workbook, sheet_name)
                except Exception:
                    continue
                columns = [str(c) for c in data.columns]
                month_col = _match_column(columns, ["지출월"])
                week_col = _match_column(columns, ["주차"])
                voucher_no_col = _match_column(columns, ["입금전표", "지출 결의 번호", "지출결의번호", "입금전표&"])
                code_col = _match_column(columns, ["계정코드"])
                category_col = _match_column(columns, ["대분류", "대분류 계정", "대분류계정"])
                desc_col = _match_column(columns, ["적요"])
                income_col = _match_column(columns, ["수입금액", "수입"])
                expense_col = _match_column(columns, ["지출금액", "지출"])
                receiver_col = _match_column(columns, ["수령자"])
                note_col = _match_column(columns, ["비고"])

                if not desc_col:
                    continue

                for _, row in data.iterrows():
                    month_str = normalize(row.get(month_col)) if month_col else ""
                    week_str = normalize(row.get(week_col)) if week_col else ""

                    # Extract month number
                    month_num = None
                    for k, v in MONTH_MAP.items():
                        if k in month_str:
                            month_num = v
                            break
                    if month_num is None:
                        # Try numeric extraction
                        m = re.search(r"(\d+)", month_str)
                        if m:
                            month_num = int(m.group(1))
                    if month_num is None or month_num < 1 or month_num > 12:
                        continue

                    # Extract week number
                    week_num = None
                    for k, v in WEEK_MAP.items():
                        if k in week_str:
                            week_num = v
                            break
                    if week_num is None:
                        m = re.search(r"(\d+)", week_str)
                        if m:
                            week_num = int(m.group(1))

                    description = normalize(row.get(desc_col)) if desc_col else ""
                    if not description:
                        continue

                    income_val = 0
                    expense_val = 0
                    try:
                        income_val = float(row.get(income_col) or 0) if income_col else 0
                        if pd.isna(income_val):
                            income_val = 0
                    except (ValueError, TypeError):
                        income_val = 0
                    try:
                        expense_val = float(row.get(expense_col) or 0) if expense_col else 0
                        if pd.isna(expense_val):
                            expense_val = 0
                    except (ValueError, TypeError):
                        expense_val = 0

                    if income_val == 0 and expense_val == 0:
                        continue

                    # The amounts in the ledger are in 천원 (thousands)
                    # Convert to actual won
                    amount_in_won = (income_val - expense_val) * 1000
                    entry_type = "income" if income_val > 0 else "expense"

                    # Build approximate voucher date
                    # Use middle of week as approximate date
                    day_approx = min(((week_num or 1) - 1) * 7 + 4, 28) if week_num else 15
                    voucher_date_str = f"{ledger_year}-{month_num:02d}-{day_approx:02d}"

                    # Generate voucher number
                    voucher_no_raw = normalize(row.get(voucher_no_col)) if voucher_no_col else ""
                    if not voucher_no_raw or voucher_no_raw.lower() == "nan":
                        voucher_no_raw = f"L-{ledger_year}{month_num:02d}-{vouchers_created + 1:05d}"

                    # Make voucher_no unique
                    voucher_no = f"{sheet_name[:4]}-{voucher_no_raw}-{vouchers_created}"

                    # Account code
                    acct_code = normalize(row.get(code_col)) if code_col else ""
                    # Clean numeric codes (e.g. "11000.0" -> "11000")
                    if acct_code:
                        try:
                            acct_code = str(int(float(acct_code)))
                        except (ValueError, TypeError):
                            pass

                    category_name = normalize(row.get(category_col)) if category_col else ""
                    receiver = normalize(row.get(receiver_col)) if receiver_col else ""
                    note_val = normalize(row.get(note_col)) if note_col else ""

                    # Find account
                    account = None
                    if acct_code:
                        account = db.scalar(select(Account).where(Account.code == acct_code))

                    # Find fund
                    fund = db.scalar(select(Fund).where(Fund.name == sheet_name))

                    try:
                        v_date = date_type.fromisoformat(voucher_date_str)
                    except Exception:
                        continue

                    db.add(
                        Voucher(
                            voucher_no=voucher_no,
                            voucher_date=v_date,
                            entry_type=entry_type,
                            description=description,
                            amount=abs(amount_in_won),
                            fund_id=fund.id if fund else None,
                            fund_name=sheet_name,
                            account_id=account.id if account else None,
                            counterparty=receiver or None,
                            note=f"{category_name} | {note_val}".strip(" |") if (category_name or note_val) else None,
                            source_workbook=ledger_workbook.name,
                            source_sheet=sheet_name,
                        )
                    )
                    vouchers_created += 1

    # ---- Seed weekly offering vouchers from 헌금현황 ----
    weekly_vouchers_created = 0
    if cash_workbook:
        year_match = re.search(r"(\d{4})년", cash_workbook.name)
        cash_year = int(year_match.group(1)) if year_match else 2025
        existing_offering = db.scalar(
            select(Voucher).where(
                Voucher.source_workbook == "weekly_offering_ui",
                Voucher.voucher_date >= date_type(cash_year, 1, 1),
                Voucher.voucher_date <= date_type(cash_year, 12, 31)
            ).limit(1)
        )
        if not existing_offering:
            excel = pd.ExcelFile(cash_workbook)
            # Find the "전체 누계" sheet
            cumulative_sheet = None
            for sn in excel.sheet_names:
                if "전체" in sn and "누계" in sn:
                    cumulative_sheet = sn
                    break
            if not cumulative_sheet:
                # Find weekly sheets (주별 헌금 장부기록)
                for sn in excel.sheet_names:
                    if "주별" in sn and "헌금" in sn:
                        cumulative_sheet = sn  # Use first weekly sheet as fallback
                        break

            if cumulative_sheet:
                try:
                    data, _ = load_sheet_frame(cash_workbook, cumulative_sheet)
                except Exception:
                    data = pd.DataFrame()

                if not data.empty:
                    columns = [str(c) for c in data.columns]
                    date_col = _match_column(columns, ["날짜", "헌금일"])
                    month_col = _match_column(columns, ["월"])
                    envelope_col = _match_column(columns, ["번호", "봉투"])
                    name_col = _match_column(columns, ["이름", "성명"])
                    dept_col = _match_column(columns, ["회별", "부서"])
                    district_col = _match_column(columns, ["구역"])
                    transfer_col = _match_column(columns, ["이체"])

                    # Map offering columns
                    offering_codes = {
                        "십일조": "11000", "주일헌금": "11200", "세계선교\n분담금": "12100",
                        "후원회비": "13000", "집회헌금": "11400", "감사헌금": "11100",
                        "기타헌금": "11500", "건축헌금": "11300", "선교회비": "12000",
                        "세계선교헌금": "12200", "사랑의헌금": "14000", "기타수입": "23000",
                    }
                    # Match offering columns
                    offering_col_map: dict[str, str] = {}
                    for col in columns:
                        col_lower = normalize(col).replace(" ", "").replace("\n", "")
                        for label, code in offering_codes.items():
                            label_clean = label.replace("\n", "")
                            if label_clean in col_lower or code in col_lower:
                                offering_col_map[code] = col
                                break

                    for _, row in data.iterrows():
                        date_val = row.get(date_col) if date_col else None
                        if date_val is None or (isinstance(date_val, str) and not date_val.strip()):
                            continue

                        name_val = normalize(row.get(name_col)) if name_col else ""
                        if not name_val:
                            continue

                        # Parse date
                        if isinstance(date_val, dt_type):
                            v_date = date_val.date()
                        elif isinstance(date_val, date_type):
                            v_date = date_val
                        elif isinstance(date_val, str):
                            try:
                                v_date = date_type.fromisoformat(date_val.split(" ")[0])
                            except Exception:
                                continue
                        else:
                            continue

                        envelope_no = normalize(row.get(envelope_col)) if envelope_col else ""
                        dept_name = normalize(row.get(dept_col)) if dept_col else ""
                        district_name = normalize(row.get(district_col)) if district_col else ""
                        is_transfer = normalize(row.get(transfer_col)).strip() if transfer_col else ""

                        # Find member
                        member = None
                        if envelope_no:
                            clean_no = re.sub(r"[^0-9]", "", envelope_no)
                            if clean_no:
                                member = db.scalar(
                                    select(Member).where(
                                        Member.year == v_date.year,
                                        Member.member_no == clean_no,
                                    )
                                )
                        if not member and name_val:
                            member = db.scalar(
                                select(Member).where(Member.name == name_val).limit(1)
                            )

                        # Create vouchers for each offering type that has value
                        for code, col_name in offering_col_map.items():
                            try:
                                val = float(row.get(col_name) or 0)
                                if pd.isna(val):
                                    val = 0
                            except (ValueError, TypeError):
                                val = 0

                            if val <= 0:
                                continue

                            # Amounts in 헌금현황 are in 천원 (thousands)
                            amount_won = val * 1000

                            account = db.scalar(select(Account).where(Account.code == code))

                            voucher_no = f"W-{v_date.isoformat()}-{envelope_no or name_val}-{code}-{weekly_vouchers_created}"

                            parts = []
                            if envelope_no:
                                parts.append(f"봉투번호 {envelope_no}")
                            if dept_name:
                                parts.append(f"회별 {dept_name}")
                            if district_name:
                                parts.append(f"구역 {district_name}")
                            if is_transfer:
                                parts.append("이체헌금")
                            composed_note = " | ".join(parts) if parts else None

                            db.add(
                                Voucher(
                                    voucher_no=voucher_no,
                                    voucher_date=v_date,
                                    entry_type="income",
                                    description=f"{v_date.month}월 주간 헌금",
                                    amount=abs(amount_won),
                                    fund_name="일반계정" if code not in ("11300", "14000", "12000", "12200") else None,
                                    account_id=account.id if account else None,
                                    member_id=member.id if member else None,
                                    counterparty=name_val,
                                    note=composed_note,
                                    source_workbook="weekly_offering_ui",
                                    source_sheet="weekly_offering_entry",
                                )
                            )
                            weekly_vouchers_created += 1

    db.commit()
    return {
        "import_batches": imported,
        "funds": funds_created,
        "accounts": accounts_created,
        "accounts_updated": accounts_updated,
        "members": members_created,
        "vouchers_from_ledger": vouchers_created,
        "vouchers_from_offerings": weekly_vouchers_created,
    }

