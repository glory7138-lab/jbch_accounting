from io import BytesIO
from decimal import Decimal

import pandas as pd
from sqlalchemy import select
from sqlalchemy.orm import Session, joinedload
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from app.models import Voucher
from app.services.dashboard_service import build_offerings_dashboard


EXPORT_COLUMNS = [
    "전표번호",
    "거래일자",
    "유형",
    "계정과목",
    "기금",
    "적요",
    "금액",
    "거래처",
    "비고",
]


def build_voucher_rows(db: Session) -> list[dict]:
    vouchers = db.scalars(
        select(Voucher)
        .options(joinedload(Voucher.account), joinedload(Voucher.fund))
        .order_by(Voucher.voucher_date.asc(), Voucher.id.asc())
    ).all()
    rows = []
    for voucher in vouchers:
        rows.append(
            {
                "전표번호": voucher.voucher_no,
                "거래일자": voucher.voucher_date.isoformat(),
                "유형": "수입" if voucher.entry_type == "income" else "지출",
                "계정과목": voucher.account.name if voucher.account else "",
                "기금": voucher.fund.name if voucher.fund else (voucher.fund_name or ""),
                "적요": voucher.description,
                "금액": float(voucher.amount),
                "거래처": voucher.counterparty or "",
                "비고": voucher.note or "",
            }
        )
    return rows


def export_vouchers_to_excel(db: Session) -> bytes:
    rows = build_voucher_rows(db)
    df = pd.DataFrame(rows, columns=EXPORT_COLUMNS)
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="전표내역")
    return buffer.getvalue()


def export_vouchers_to_markdown(db: Session) -> str:
    rows = build_voucher_rows(db)
    if not rows:
        return "# 전표 내역\n\n등록된 전표가 없습니다.\n"
    df = pd.DataFrame(rows, columns=EXPORT_COLUMNS)
    return "# 전표 내역\n\n" + df.to_markdown(index=False)


def export_offerings_dashboard_to_excel(
    db: Session,
    start_ym: str | None = None,
    end_ym: str | None = None,
    department: str | None = None,
    account_id: int | None = None,
) -> bytes:
    data = build_offerings_dashboard(db, start_ym, end_ym, department, account_id)
    wb = Workbook()
    
    ws1 = wb.active
    ws1.title = "헌금 요약 및 분석"
    ws1.views.sheetView[0].showGridLines = True
    
    font_title = Font(name="맑은 고딕", size=14, bold=True, color="1B365D")
    font_subtitle = Font(name="맑은 고딕", size=9, italic=True, color="555555")
    font_section = Font(name="맑은 고딕", size=11, bold=True, color="1B365D")
    font_header = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
    font_data = Font(name="맑은 고딕", size=10)
    font_data_bold = Font(name="맑은 고딕", size=10, bold=True)
    
    fill_header = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
    fill_summary_label = PatternFill(start_color="F2F4F7", end_color="F2F4F7", fill_type="solid")
    fill_summary_value = PatternFill(start_color="E6EEFA", end_color="E6EEFA", fill_type="solid")
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    
    thin_side = Side(border_style="thin", color="CCCCCC")
    double_side = Side(border_style="double", color="1B365D")
    thick_bottom = Side(border_style="medium", color="1B365D")
    
    border_thin = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    border_summary = Border(left=thin_side, right=thin_side, top=thin_side, bottom=double_side)
    border_header = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thick_bottom)
    
    ws1["A1"] = "헌금 통계 분석 보고서"
    ws1["A1"].font = font_title
    
    period_str = f"{start_ym or '전체'} ~ {end_ym or '전체'}"
    filter_desc = f"기간: {period_str} | 회별(부서): {department or '전체'} | 헌금종류: {'지정됨' if account_id else '전체'}"
    ws1["A2"] = filter_desc
    ws1["A2"].font = font_subtitle
    
    summary_cards = [
        ("총 헌금액", data["total_amount"], "₩#,##0"),
        ("총 헌금 횟수", data["total_count"], "#,##0건"),
        ("참여 인원수", data["unique_participants"], "#,##0명"),
        ("1인당 평균 헌금액", data["average_amount_per_person"], "₩#,##0")
    ]
    
    for idx, (label, val, fmt) in enumerate(summary_cards):
        col_letter = get_column_letter(idx + 1)
        cell_lbl = ws1[f"{col_letter}4"]
        cell_lbl.value = label
        cell_lbl.font = font_data_bold
        cell_lbl.fill = fill_summary_label
        cell_lbl.alignment = align_center
        cell_lbl.border = border_thin
        
        cell_val = ws1[f"{col_letter}5"]
        cell_val.value = val
        cell_val.font = font_title
        cell_val.fill = fill_summary_value
        cell_val.alignment = align_center
        cell_val.border = border_summary
        cell_val.number_format = fmt
        
    ws1["A7"] = "1. 헌금 종류별 집계"
    ws1["A7"].font = font_section
    
    headers_acc = ["헌금 종류", "총 헌금액", "건수", "비중(%)"]
    for col_idx, h in enumerate(headers_acc, start=1):
        cell = ws1.cell(row=8, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_header
        
    row_idx = 9
    for acc in data["by_account"]:
        ws1.cell(row=row_idx, column=1, value=acc["account_name"]).alignment = align_left
        ws1.cell(row=row_idx, column=2, value=acc["total_amount"]).number_format = "₩#,##0"
        ws1.cell(row=row_idx, column=3, value=acc["total_count"]).number_format = "#,##0"
        ws1.cell(row=row_idx, column=4, value=acc["percentage"] / 100).number_format = "0.0%"
        
        for c in range(1, 5):
            cell = ws1.cell(row=row_idx, column=c)
            cell.font = font_data
            cell.border = border_thin
            if c in (2, 3, 4):
                cell.alignment = align_right
        row_idx += 1
        
    ws1.cell(row=row_idx, column=1, value="합계").font = font_data_bold
    ws1.cell(row=row_idx, column=1).alignment = align_center
    ws1.cell(row=row_idx, column=2, value=data["total_amount"]).number_format = "₩#,##0"
    ws1.cell(row=row_idx, column=3, value=data["total_count"]).number_format = "#,##0"
    ws1.cell(row=row_idx, column=4, value=1.0).number_format = "0.0%"
    for c in range(1, 5):
        cell = ws1.cell(row=row_idx, column=c)
        cell.font = font_data_bold
        cell.border = border_summary
        if c in (2, 3, 4):
            cell.alignment = align_right
            
    row_idx += 3
    ws1.cell(row=row_idx, column=1, value="2. 회별(부서별) 집계").font = font_section
    row_idx += 1
    
    headers_dept = ["회별(부서)", "총 헌금액", "건수", "비중(%)"]
    for col_idx, h in enumerate(headers_dept, start=1):
        cell = ws1.cell(row=row_idx, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_header
        
    row_idx += 1
    for dept in data["by_department"]:
        ws1.cell(row=row_idx, column=1, value=dept["department_name"]).alignment = align_left
        ws1.cell(row=row_idx, column=2, value=dept["total_amount"]).number_format = "₩#,##0"
        ws1.cell(row=row_idx, column=3, value=dept["total_count"]).number_format = "#,##0"
        ws1.cell(row=row_idx, column=4, value=dept["percentage"] / 100).number_format = "0.0%"
        
        for c in range(1, 5):
            cell = ws1.cell(row=row_idx, column=c)
            cell.font = font_data
            cell.border = border_thin
            if c in (2, 3, 4):
                cell.alignment = align_right
        row_idx += 1
        
    ws1.cell(row=row_idx, column=1, value="합계").font = font_data_bold
    ws1.cell(row=row_idx, column=1).alignment = align_center
    ws1.cell(row=row_idx, column=2, value=data["total_amount"]).number_format = "₩#,##0"
    ws1.cell(row=row_idx, column=3, value=data["total_count"]).number_format = "#,##0"
    ws1.cell(row=row_idx, column=4, value=1.0).number_format = "0.0%"
    for c in range(1, 5):
        cell = ws1.cell(row=row_idx, column=c)
        cell.font = font_data_bold
        cell.border = border_summary
        if c in (2, 3, 4):
            cell.alignment = align_right
            
    ws2 = wb.create_sheet(title="기간별 추이")
    ws2.views.sheetView[0].showGridLines = True
    
    ws2["A1"] = "기간별 헌금 추이 분석"
    ws2["A1"].font = font_title
    
    ws2["A3"] = "1. 월별 헌금 추이"
    ws2["A3"].font = font_section
    
    acc_set = set()
    for m in data["monthly_trends"]:
        acc_set.update(m["amounts"].keys())
    acc_list = sorted(list(acc_set))
    
    headers_trend = ["연월", "총 헌금액", "참여자 수", "건수"] + acc_list
    for col_idx, h in enumerate(headers_trend, start=1):
        cell = ws2.cell(row=4, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_header
        
    row_idx = 5
    for m in data["monthly_trends"]:
        ws2.cell(row=row_idx, column=1, value=m["period"]).alignment = align_center
        ws2.cell(row=row_idx, column=2, value=m["total_amount"]).number_format = "₩#,##0"
        ws2.cell(row=row_idx, column=3, value=m["total_participants"]).number_format = "#,##0"
        ws2.cell(row=row_idx, column=4, value=m["total_count"]).number_format = "#,##0"
        
        for a_idx, acc_name in enumerate(acc_list, start=5):
            val = m["amounts"].get(acc_name, Decimal(0))
            ws2.cell(row=row_idx, column=a_idx, value=float(val))
            ws2.cell(row=row_idx, column=a_idx).number_format = "₩#,##0"
            
        for c in range(1, len(headers_trend) + 1):
            cell = ws2.cell(row=row_idx, column=c)
            cell.font = font_data
            cell.border = border_thin
            if c >= 2:
                cell.alignment = align_right
        row_idx += 1
        
    row_idx += 3
    ws2.cell(row=row_idx, column=1, value="2. 연도별 헌금 추이").font = font_section
    row_idx += 1
    
    acc_set_y = set()
    for y in data["yearly_trends"]:
        acc_set_y.update(y["amounts"].keys())
    acc_list_y = sorted(list(acc_set_y))
    
    headers_trend_y = ["연도", "총 헌금액", "참여자 수", "건수"] + acc_list_y
    for col_idx, h in enumerate(headers_trend_y, start=1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_header
        
    row_idx += 1
    for y in data["yearly_trends"]:
        ws2.cell(row=row_idx, column=1, value=y["period"]).alignment = align_center
        ws2.cell(row=row_idx, column=2, value=y["total_amount"]).number_format = "₩#,##0"
        ws2.cell(row=row_idx, column=3, value=y["total_participants"]).number_format = "#,##0"
        ws2.cell(row=row_idx, column=4, value=y["total_count"]).number_format = "#,##0"
        
        for a_idx, acc_name in enumerate(acc_list_y, start=5):
            val = y["amounts"].get(acc_name, Decimal(0))
            ws2.cell(row=row_idx, column=a_idx, value=float(val))
            ws2.cell(row=row_idx, column=a_idx).number_format = "₩#,##0"
            
        for c in range(1, len(headers_trend_y) + 1):
            cell = ws2.cell(row=row_idx, column=c)
            cell.font = font_data
            cell.border = border_thin
            if c >= 2:
                cell.alignment = align_right
        row_idx += 1
        
    ws3 = wb.create_sheet(title="금액구간 분포")
    ws3.views.sheetView[0].showGridLines = True
    
    ws3["A1"] = "헌금 규모별 분포 분석"
    ws3["A1"].font = font_title
    
    ws3["A3"] = "1. 헌금 1회당 금액 구간별 통계"
    ws3["A3"].font = font_section
    
    headers_range = ["금액 구간", "헌금 횟수(건수)", "총 헌금액", "건수 비중(%)"]
    for col_idx, h in enumerate(headers_range, start=1):
        cell = ws3.cell(row=4, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_header
        
    row_idx = 5
    for item in data["by_amount_range"]:
        ws3.cell(row=row_idx, column=1, value=item["range_label"]).alignment = align_left
        ws3.cell(row=row_idx, column=2, value=item["total_count"]).number_format = "#,##0"
        ws3.cell(row=row_idx, column=3, value=item["total_amount"]).number_format = "₩#,##0"
        
        pct = (item["total_count"] / data["total_count"]) if data["total_count"] > 0 else 0
        ws3.cell(row=row_idx, column=4, value=pct).number_format = "0.0%"
        
        for c in range(1, 5):
            cell = ws3.cell(row=row_idx, column=c)
            cell.font = font_data
            cell.border = border_thin
            if c in (2, 3, 4):
                cell.alignment = align_right
        row_idx += 1
        
    ws3.cell(row=row_idx, column=1, value="합계").font = font_data_bold
    ws3.cell(row=row_idx, column=1).alignment = align_center
    ws3.cell(row=row_idx, column=2, value=data["total_count"]).number_format = "#,##0"
    ws3.cell(row=row_idx, column=3, value=data["total_amount"]).number_format = "₩#,##0"
    ws3.cell(row=row_idx, column=4, value=1.0).number_format = "0.0%"
    for c in range(1, 5):
        cell = ws3.cell(row=row_idx, column=c)
        cell.font = font_data_bold
        cell.border = border_summary
        if c in (2, 3, 4):
            cell.alignment = align_right
            
    for ws in (ws1, ws2, ws3):
        for col in ws.columns:
            max_len = 0
            for cell in col:
                val = str(cell.value or '')
                val_len = len(val.encode('utf-8'))
                if val_len > max_len:
                    max_len = val_len
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len // 2 + 4, 12)
            
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()

